# -*- coding: utf-8 -*-
"""
Excel文件适配器

支持 .xls、.xlsx 和 .ods 文件的解析。
"""

from pathlib import Path

from .base_adapter import BaseFileAdapter


class ExcelFileAdapter(BaseFileAdapter):
    """Excel文件适配器"""

    def __init__(self, file_path: str, max_file_size: int = 30 * 1024 * 1024):
        super().__init__(file_path, max_file_size)
        self._file_content: str | None = None
        self._workbook: dict | None = None

    def _get_file_description(self) -> str:
        """获取文件描述"""
        return "Excel File"

    async def _load_workbook(self) -> dict | None:
        """加载工作簿"""
        if self._workbook is not None:
            return self._workbook

        from aiofiles import open as aio_open
        import aiofiles

        stat = await aiofiles.os.stat(self.file_path)
        if stat.st_size > self.max_file_size:
            return None

        try:
            async with aio_open(self.file_path, "rb") as f:
                buffer = await f.read()

            # 根据文件扩展名选择解析方式
            ext = Path(self.file_path).suffix.lower()

            if ext == ".xls":
                # 使用 xlrd 解析旧版 .xls 文件
                return await self._parse_xls(buffer)
            else:
                # 使用 openpyxl 解析 .xlsx 文件
                return await self._parse_xlsx(buffer)

        except Exception as e:
            from app.core.logger import log
            log.error(f"Error parsing Excel file: {e}")
            return None

    async def _parse_xlsx(self, buffer: bytes) -> dict | None:
        """解析 .xlsx 文件"""
        from openpyxl import load_workbook
        from io import BytesIO

        try:
            wb = load_workbook(BytesIO(buffer), data_only=True)
        except Exception as e:
            # 如果 data_only=True 失败，尝试不使用 data_only
            try:
                wb = load_workbook(BytesIO(buffer), data_only=False)
            except Exception:
                raise e

        # 提取所有工作表数据
        sheets_data = {}
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            rows = []
            for row in sheet.iter_rows():
                row_data = []
                for cell in row:
                    value = cell.value
                    if value is None:
                        row_data.append("")
                    else:
                        row_data.append(str(value))
                rows.append(row_data)

            sheets_data[sheet_name] = {
                "rows": rows,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
            }

        self._workbook = {
            "sheet_names": wb.sheetnames,
            "sheets": sheets_data,
        }

        return self._workbook

    async def _parse_xls(self, buffer: bytes) -> dict | None:
        """解析 .xls 文件"""
        try:
            import xlrd
            from io import BytesIO

            book = xlrd.open_workbook(file_contents=buffer)

            sheets_data = {}
            sheet_names = book.sheet_names()

            for sheet_idx in range(book.nsheets):
                sheet = book.sheet_by_index(sheet_idx)
                sheet_name = sheet_names[sheet_idx]

                rows = []
                for row_idx in range(sheet.nrows):
                    row_data = []
                    for col_idx in range(sheet.ncols):
                        cell = sheet.cell(row_idx, col_idx)
                        value = cell.value
                        if value is None:
                            row_data.append("")
                        else:
                            row_data.append(str(value))
                    rows.append(row_data)

                sheets_data[sheet_name] = {
                    "rows": rows,
                    "max_row": sheet.nrows,
                    "max_column": sheet.ncols,
                }

            self._workbook = {
                "sheet_names": sheet_names,
                "sheets": sheets_data,
            }

            return self._workbook

        except ImportError:
            from app.core.logger import log
            log.warning("xlrd not installed, falling back to pandas for .xls files")
            return await self._parse_with_pandas(buffer)

    async def _parse_with_pandas(self, buffer: bytes) -> dict | None:
        """使用 pandas 解析 Excel 文件（备用方案）"""
        import pandas as pd
        from io import BytesIO

        try:
            # 读取所有工作表
            xls = pd.ExcelFile(BytesIO(buffer))

            sheets_data = {}
            for sheet_name in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet_name)

                # 将 DataFrame 转换为列表
                rows = []
                # 添加表头
                headers = [str(col) for col in df.columns.tolist()]
                rows.append(headers)

                # 添加数据行
                for _, row in df.iterrows():
                    row_data = [str(val) if val is not None else "" for val in row.tolist()]
                    rows.append(row_data)

                sheets_data[sheet_name] = {
                    "rows": rows,
                    "max_row": len(rows),
                    "max_column": len(headers) if headers else 0,
                }

            self._workbook = {
                "sheet_names": xls.sheet_names,
                "sheets": sheets_data,
            }

            return self._workbook

        except Exception as e:
            from app.core.logger import log
            log.error(f"Error parsing Excel with pandas: {e}")
            return None

    def _generate_sheet_content(self, sheet_data: dict) -> str:
        """生成工作表内容"""
        rows = sheet_data.get("rows", [])
        if not rows:
            return ""

        # 转换为制表符分隔的格式
        return "\n".join(["\t".join(row) for row in rows])

    async def get_content(self) -> str | None:
        """获取文件原始内容"""
        if self._file_content is None:
            workbook = await self._load_workbook()
            if workbook:
                sheets_content = []
                for sheet_name, sheet_data in workbook["sheets"].items():
                    content = self._generate_sheet_content(sheet_data)
                    sheets_content.append(f"Sheet: {sheet_name}\n{content}")
                self._file_content = "\n\n".join(sheets_content)

        return self._file_content

    async def get_llm_content(self) -> str | None:
        """获取适合LLM处理的内容格式"""
        workbook = await self._load_workbook()
        if not workbook:
            return None

        file_description = f"""# Excel File Description

## Basic Excel File Information
* **Total Sheets:** {len(workbook["sheet_names"])}

## Sheets Information
"""

        # 为每个工作表生成预览
        sheets_content = []
        for sheet_name in workbook["sheet_names"]:
            sheet_data = workbook["sheets"][sheet_name]
            total_rows = sheet_data.get("max_row", 0)
            total_cols = sheet_data.get("max_column", 0)

            # 获取前10行数据
            rows = sheet_data.get("rows", [])
            preview_rows = rows[:10]
            preview_content = "\n".join(["\t".join(row) for row in preview_rows])

            sheets_content.append(f"""### Sheet: {sheet_name}
* **Total Rows:** {total_rows}
* **Total Columns:** {total_cols}

#### Data Preview (First 10 Rows)
```
{preview_content}
```
""")

        return file_description + "\n".join(sheets_content)

    async def get_thumbnail(self) -> str | None:
        """获取缩略图"""
        return None
